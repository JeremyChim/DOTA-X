"""
-*- coding: utf-8 -*-
Time: 2023/10/30 17:09
Auth: Jeremy.Chim
File: InitConf.py
IDE: PyCharm
GitHub: https://github.com/JeremyChim/DOTA-X
"""


from openpyxl import load_workbook


def loadConf(filePath: str, sheetName: str) -> list:
    """
    Load Config Excel.
    :param filePath: ex.'config.xlsx'
    :param sheetName: ex.'hr'
    :return: listConf.
    """
    wb = load_workbook(filePath)
    ws = wb[sheetName]
    maxX = ws.max_row
    maxY = ws.max_column
    listConf = []

    for x in range(2, maxX + 1):
        dictConf = {}

        for y in range(1, maxY + 1):
            title = ws.cell(1, y).value  # title
            cellVal = ws.cell(x, y).value  # cell value
            dictConf[title] = cellVal

        listConf.append(dictConf)
    return listConf


def mainScript(strOld: str, listConf: list) -> str:
    """
    main Script.
    """
    listStrOld = strOld.split('\n')
    listStrNew = []

    for lineStrOld in listStrOld:
        flag = False

        for dictConf in listConf:
            key = dictConf['Keyword']
            mul = dictConf['Mul']
            add = dictConf['Add']
            poi = dictConf['Point']

            if key in lineStrOld:
                flag = True
                listStrOld2 = lineStrOld.split('"')
                valStrOld = listStrOld2[-2]
                valStrOld2 = valStrOld.split(' ')
                listValNew = []

                for valOld in valStrOld2:
                    # noinspection PyBroadException
                    try:
                        valNew = float(valOld) * mul + add
                        valNew = f'{valNew:.{poi}f}'
                        listValNew.append(valNew)
                    except:
                        listValNew.append(valOld)

                strValNew = ' '.join(listValNew)
                listStrOld2[-2] = strValNew
                lineStrNew = '"'.join(listStrOld2)
                listStrNew.append(lineStrNew)

        if flag is False:
            listStrNew.append(lineStrOld)

    strNew = '\n'.join(listStrNew)
    return strNew


class InitConf(object):
    """
    init Config.
    """

    def __init__(self):
        conf = 'config.xlsx'
        self.conf_hr = loadConf(conf, 'Hero')
        self.conf_ab = loadConf(conf, 'Ability')
        self.conf_un = loadConf(conf, 'Unit')

    def scriptHero(self, strOld):
        """
        Script Hero.
        """
        return mainScript(strOld, self.conf_hr)

    def scriptAbility(self, strOld):
        """
        Script Ability.
        """
        return mainScript(strOld, self.conf_ab)

    def scriptUnit(self, strOld):
        """
        Script Unit.
        """
        return mainScript(strOld, self.conf_un)


if __name__ == '__main__':

    testText = '''
            "BountyXP"					"69"		// Experience earn.
            "BountyGoldMin"				"43"		// Gold earned min.
            "BountyGoldMax"				"52"		// Gold earned max.
    '''

    print(InitConf().scriptUnit(testText))
